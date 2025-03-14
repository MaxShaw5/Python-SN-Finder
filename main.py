from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import time
import os
import sys
from variables import main_URL, username, password
from openpyxl import load_workbook


#Begin spreadsheet initialization loop
spread_sheet_path = "Path_to_SS_Containing_User_Names"
#Loading spreadsheet into Python
wb = load_workbook(spread_sheet_path)
ws = wb['Sheet1']

all_rows = list(ws.rows)

#Creating an empty list in python to store the names
names = []

for row in ws:
    name = row[0].value
    names.append(name)

#Opening Firefox
driver = webdriver.Firefox()

#Going to the desired website (in this case an internal Asset Management WebUI)
driver.get(main_URL)
driver.maximize_window()

#Sign in block
WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "LoginPanel0_Username")))
un_form = driver.find_element(By.ID, "LoginPanel0_Username")
un_form.send_keys(username)

pass_form = driver.find_element(By.ID, "LoginPanel0_Password")
pass_form.send_keys(password + Keys.ENTER)
#End Sign in block

#Clearing the prefilled search bar
WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "select2-search-choice-close")))
clear_filter = driver.find_element(By.CLASS_NAME, "select2-search-choice-close")
clear_filter.click()


#Initializing search bar variable
WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "Lanman2_Default_PurchaseOrderDetailsGrid0_QuickSearchInput")))
search_bar = driver.find_element(By.ID, "Lanman2_Default_PurchaseOrderDetailsGrid0_QuickSearchInput")

#The following 4 operations click on the sort button twice to sort for the newest hardware
sort_by = driver.find_element(By.CSS_SELECTOR, "div[id*='sleekgrid_'][id*='PurchaseOrderCreatedDate']")
sort_by.click()
time.sleep(1)
sort_by.click()


#This loop will enter a user's name from a spreadsheet, pull the name of their PC and the date in which the lease ends and add them to the same spreadsheet in the same row as the user's name
for row, name in enumerate(names, start=1):
    try:
        search_bar.send_keys(name + Keys.ENTER)
        time.sleep(1)
        sn_cell = driver.find_element(By.CLASS_NAME, "slick-cell.l0.r0")
        output = sn_cell.find_element(By.XPATH, ".//a[@class='s-EditLink s-Default-PurchaseOrderDetailsLink']")
        print(f"SN {output.text} found! Adding to spreadsheet.")
        ws[f'B{row}'] = output.text
        print(f"Appended SN # {output.text} for {name} to row {row}")
        eol_cell = driver.find_element(By.CLASS_NAME, "slick-cell.l9.r9")
        ws[f'C{row}'] = eol_cell.text
        print(f"EOL data for {name} found. Adding to spreadsheet in row {row}")
        time.sleep(1)
        search_bar.clear()
    except Exception as e:
        ws[f'B{row}'] = "Not Found"
        print(f"Error for {name}. Appended 'Not Found' to row {row}")
        search_bar.clear()
#Saves the spreadsheet for viewing
wb.save(spread_sheet_path)

print(f"Information retrieved for all users. You can find it at {spread_sheet_path}. Thank you!")

#Pause for user to read message
time.sleep(10)

driver.quit()
sys.exit()